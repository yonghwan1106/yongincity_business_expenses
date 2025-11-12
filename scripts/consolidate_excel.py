#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
용인시장 업무추진비 엑셀 데이터 통합 스크립트
2022.7.1 ~ 2025.9.30 전체 데이터를 CSV로 변환
"""

import openpyxl
import csv
import re
from datetime import datetime

# 파일 경로
EXCEL_FILE = 'C:/Users/user/yongincity_business_expenses/data/용인시장 업무추진비 집행내역(2022.7.1.~2025.9.30.).xlsx'
OUTPUT_CSV = 'C:/Users/user/yongincity_business_expenses/data/용인시장_업무추진비_전체_2022-2025.csv'

# 통합 데이터 저장
all_records = []
errors = []

def clean_value(value):
    """셀 값 정리"""
    if value is None:
        return ''
    return str(value).strip()

def is_header_row(row_data):
    """헤더 행인지 확인"""
    first_cell = clean_value(row_data[0])
    return first_cell in ['번호', '순번']

def is_summary_row(row_data):
    """요약 행인지 확인 (예: "32 건", "9,962,010" 등)"""
    first_cell = clean_value(row_data[0])
    second_cell = clean_value(row_data[1])
    # 첫 번째 셀이 비어있고, 4-5번째 셀에 "건"이 있으면 요약 행
    if not first_cell and ('건' in second_cell or '건' in clean_value(row_data[3]) or '건' in clean_value(row_data[4])):
        return True
    return False

def is_title_row(row_data):
    """제목 행인지 확인 (예: "2022년 7월 용인시장 업무추진비 집행내역")"""
    first_cell = clean_value(row_data[0])
    return '업무추진비' in first_cell and '집행내역' in first_cell

def is_data_row(row_data):
    """실제 데이터 행인지 확인"""
    first_cell = clean_value(row_data[0])
    # 번호가 숫자인지 확인
    try:
        int(first_cell)
        return True
    except ValueError:
        return False

def parse_amount(value):
    """금액 파싱 (쉼표 제거, 숫자만 추출)"""
    if value is None or value == '':
        return 0
    value_str = str(value).replace(',', '').replace('.', '')
    # 숫자만 추출
    numbers = re.findall(r'\d+', value_str)
    if numbers:
        return int(numbers[0])
    return 0

def parse_personnel(value):
    """인원수 파싱"""
    if value is None or value == '':
        return 0
    value_str = str(value).strip()
    # "용인시장" 같은 텍스트면 1로 처리
    if value_str in ['용인시장', '-']:
        return 1
    # 숫자 추출
    numbers = re.findall(r'\d+', value_str)
    if numbers:
        return int(numbers[0])
    return 0

def parse_row(row_data, has_user_column=False):
    """행 데이터 파싱"""
    try:
        if has_user_column:
            # 사용자 컬럼이 있는 경우 (최신 포맷)
            # 번호, 사용자, 사용일시, 사용장소, 집행목적, 대상인원, 사용금액, 결제방법, 비고(선택)
            record = {
                '번호': clean_value(row_data[0]),
                '사용자': clean_value(row_data[1]) if row_data[1] else '용인시장',
                '사용일시': clean_value(row_data[2]),
                '사용장소': clean_value(row_data[3]),
                '집행목적': clean_value(row_data[4]),
                '대상인원': parse_personnel(row_data[5]),
                '사용금액': parse_amount(row_data[6]),
                '결제방법': clean_value(row_data[7]),
                '비목': '',  # 나중에 분류
                '비고': clean_value(row_data[8]) if len(row_data) > 8 else ''
            }
        else:
            # 사용자 컬럼이 없는 경우 (초기 포맷)
            # 번호, 사용일시, 사용장소, 집행목적, 대상인원, 사용금액, 결제방법, 비고(선택)
            record = {
                '번호': clean_value(row_data[0]),
                '사용자': '용인시장',
                '사용일시': clean_value(row_data[1]),
                '사용장소': clean_value(row_data[2]),
                '집행목적': clean_value(row_data[3]),
                '대상인원': parse_personnel(row_data[4]),
                '사용금액': parse_amount(row_data[5]),
                '결제방법': clean_value(row_data[6]),
                '비목': '',
                '비고': clean_value(row_data[7]) if len(row_data) > 7 else ''
            }

        return record
    except Exception as e:
        raise ValueError(f"행 파싱 오류: {e}, 데이터: {row_data}")

def detect_format(sheet):
    """시트의 포맷 감지 (헤더 있는지, 사용자 컬럼 있는지)"""
    # 첫 10개 행 확인
    for row_num in range(1, min(11, sheet.max_row + 1)):
        row_data = [sheet.cell(row=row_num, column=col).value for col in range(1, 10)]

        if is_header_row(row_data):
            # 헤더 발견 - 사용자 컬럼 있는지 확인
            has_user = '사용자' in str(row_data)
            return {
                'has_header': True,
                'header_row': row_num,
                'data_start_row': row_num + 2,  # 헤더 다음에 요약 행이 있으므로 +2
                'has_user_column': has_user
            }

    # 헤더가 없으면 연속 데이터
    # 첫 번째 행이 데이터인지 확인
    first_row = [sheet.cell(row=1, column=col).value for col in range(1, 10)]
    if is_data_row(first_row):
        # 사용자 컬럼 유무는 두 번째 셀이 날짜 형식인지로 판단
        second_cell = clean_value(first_row[1])
        has_user = '용인시장' in second_cell or (not re.search(r'\d{4}[-/]\d{2}', second_cell))

        return {
            'has_header': False,
            'header_row': None,
            'data_start_row': 1,
            'has_user_column': has_user
        }

    return None

print("=" * 80)
print("용인시장 업무추진비 데이터 통합 시작")
print("=" * 80)

# 엑셀 파일 로드
print(f"\n엑셀 파일 로드 중: {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE)
print(f"총 {len(wb.sheetnames)}개 시트 발견")

# 각 시트 처리
for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
    sheet = wb[sheet_name]
    print(f"\n[{sheet_idx}/{len(wb.sheetnames)}] {sheet_name} 처리 중... (총 {sheet.max_row}행)")

    # 포맷 감지
    format_info = detect_format(sheet)
    if not format_info:
        print(f"  [WARNING] 포맷 감지 실패 - 건너뜀")
        errors.append(f"{sheet_name}: 포맷 감지 실패")
        continue

    print(f"  - 헤더: {'있음' if format_info['has_header'] else '없음(연속 데이터)'}")
    print(f"  - 사용자 컬럼: {'있음' if format_info['has_user_column'] else '없음'}")
    print(f"  - 데이터 시작 행: {format_info['data_start_row']}")

    # 데이터 행 처리
    sheet_records = 0
    for row_num in range(format_info['data_start_row'], sheet.max_row + 1):
        row_data = [sheet.cell(row=row_num, column=col).value for col in range(1, 10)]

        # 빈 행 건너뛰기
        if all(cell is None or str(cell).strip() == '' for cell in row_data):
            continue

        # 데이터 행인지 확인
        if not is_data_row(row_data):
            continue

        try:
            record = parse_row(row_data, format_info['has_user_column'])

            # 유효성 검증
            if record['사용금액'] <= 0:
                print(f"  [WARNING] 행 {row_num}: 금액이 0 이하 - 건너뜀")
                continue

            if not record['사용일시']:
                print(f"  [WARNING] 행 {row_num}: 사용일시 누락 - 건너뜀")
                continue

            all_records.append(record)
            sheet_records += 1

        except Exception as e:
            print(f"  [ERROR] 행 {row_num} 파싱 오류: {e}")
            errors.append(f"{sheet_name} 행 {row_num}: {e}")

    print(f"  [OK] {sheet_records}건 추가됨")

print(f"\n{'=' * 80}")
print(f"통합 완료: 총 {len(all_records)}건의 레코드")
print(f"{'=' * 80}")

# CSV 파일로 저장
print(f"\nCSV 파일 저장 중: {OUTPUT_CSV}")
with open(OUTPUT_CSV, 'w', encoding='utf-8-sig', newline='') as f:
    fieldnames = ['번호', '사용자', '사용일시', '사용장소', '집행목적', '대상인원', '사용금액', '결제방법', '비목', '비고']
    writer = csv.DictWriter(f, fieldnames=fieldnames)

    writer.writeheader()
    for record in all_records:
        writer.writerow(record)

print(f"[OK] CSV 파일 저장 완료")

# 통계 출력
total_amount = sum(r['사용금액'] for r in all_records)
print(f"\n{'=' * 80}")
print("통합 데이터 통계")
print(f"{'=' * 80}")
print(f"총 건수: {len(all_records):,}건")
print(f"총 금액: {total_amount:,}원 ({total_amount/100000000:.2f}억원)")
print(f"평균 금액: {total_amount/len(all_records):,.0f}원")

# 에러 리포트
if errors:
    print(f"\n[WARNING]  발생한 오류 ({len(errors)}건):")
    for error in errors[:10]:  # 최대 10개만 출력
        print(f"  - {error}")
    if len(errors) > 10:
        print(f"  ... 외 {len(errors) - 10}건")

print(f"\n{'=' * 80}")
print("완료!")
print(f"{'=' * 80}")
