#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
용인시장 업무추진비 엑셀 데이터 통합 스크립트 V3
Sheet28(2024년 1월)부터 비목 컬럼 포함
"""

import openpyxl
import csv
import re

# 파일 경로
EXCEL_FILE = 'C:/Users/user/yongincity_business_expenses/data/용인시장 업무추진비 집행내역(2022.7.1.~2025.9.30.).xlsx'
OUTPUT_CSV = 'C:/Users/user/yongincity_business_expenses/data/용인시장_업무추진비_전체_2022-2025.csv'

# Sheet28부터 비목 컬럼 있음
SHEET_WITH_CATEGORY_START = 28

# 통합 데이터 저장
all_records = []
errors = []

def clean_value(value):
    """셀 값 정리"""
    if value is None:
        return ''
    return str(value).strip()

def is_numeric(value):
    """숫자인지 확인"""
    if value is None or value == '':
        return False
    try:
        int(str(value).strip())
        return True
    except ValueError:
        return False

def parse_amount(value):
    """금액 파싱"""
    if value is None or value == '':
        return 0
    value_str = str(value).replace(',', '').replace('.', '').replace(' ', '')
    numbers = re.findall(r'\d+', value_str)
    if numbers:
        return int(numbers[0])
    return 0

def parse_personnel(value):
    """인원수 파싱"""
    if value is None or value == '':
        return 0
    value_str = str(value).strip()
    if value_str in ['-', '용인시장']:
        return 1
    numbers = re.findall(r'\d+', value_str)
    if numbers:
        return int(numbers[0])
    return 0

def detect_data_start(sheet):
    """데이터가 시작하는 행 찾기"""
    for row_num in range(1, min(20, sheet.max_row + 1)):
        first_cell = clean_value(sheet.cell(row=row_num, column=1).value)
        second_cell = clean_value(sheet.cell(row=row_num, column=2).value)

        if first_cell == '1':
            if '용인시장' in second_cell:
                return {'start_row': row_num, 'has_user_column': True}
            elif re.search(r'\d{4}[-/]\d{2}', second_cell) or ':' in second_cell:
                return {'start_row': row_num, 'has_user_column': False}

    first_row_first_cell = clean_value(sheet.cell(row=1, column=1).value)
    first_row_second_cell = clean_value(sheet.cell(row=1, column=2).value)

    if is_numeric(first_row_first_cell):
        if '용인시장' in first_row_second_cell:
            return {'start_row': 1, 'has_user_column': True}
        elif re.search(r'\d{4}[-/]\d{2}', first_row_second_cell) or ':' in first_row_second_cell:
            return {'start_row': 1, 'has_user_column': False}

    second_row_first_cell = clean_value(sheet.cell(row=2, column=1).value)
    second_row_second_cell = clean_value(sheet.cell(row=2, column=2).value)

    if is_numeric(second_row_first_cell):
        if '용인시장' in second_row_second_cell:
            return {'start_row': 2, 'has_user_column': True}
        elif re.search(r'\d{4}[-/]\d{2}', second_row_second_cell) or ':' in second_row_second_cell:
            return {'start_row': 2, 'has_user_column': False}

    return None

def parse_row(sheet, row_num, has_user_column, has_category):
    """행 데이터 파싱"""
    try:
        if has_user_column:
            # Sheet28 이후: 번호, 사용자, 사용일시, 사용장소, 집행목적, 대상인원, 사용금액, 결제방법, 비목, 비고
            record = {
                '번호': clean_value(sheet.cell(row=row_num, column=1).value),
                '사용자': clean_value(sheet.cell(row=row_num, column=2).value) or '용인시장',
                '사용일시': clean_value(sheet.cell(row=row_num, column=3).value),
                '사용장소': clean_value(sheet.cell(row=row_num, column=4).value),
                '집행목적': clean_value(sheet.cell(row=row_num, column=5).value),
                '대상인원': parse_personnel(sheet.cell(row=row_num, column=6).value),
                '사용금액': parse_amount(sheet.cell(row=row_num, column=7).value),
                '결제방법': clean_value(sheet.cell(row=row_num, column=8).value),
                '비목': clean_value(sheet.cell(row=row_num, column=9).value) if has_category else '',
                '비고': clean_value(sheet.cell(row=row_num, column=10).value) if has_category and sheet.max_column >= 10 else clean_value(sheet.cell(row=row_num, column=9).value) if not has_category and sheet.max_column >= 9 else ''
            }
        else:
            # 사용자 컬럼 없음: 번호, 사용일시, 사용장소, 집행목적, 대상인원, 사용금액, 결제방법, 비고
            record = {
                '번호': clean_value(sheet.cell(row=row_num, column=1).value),
                '사용자': '용인시장',
                '사용일시': clean_value(sheet.cell(row=row_num, column=2).value),
                '사용장소': clean_value(sheet.cell(row=row_num, column=3).value),
                '집행목적': clean_value(sheet.cell(row=row_num, column=4).value),
                '대상인원': parse_personnel(sheet.cell(row=row_num, column=5).value),
                '사용금액': parse_amount(sheet.cell(row=row_num, column=6).value),
                '결제방법': clean_value(sheet.cell(row=row_num, column=7).value),
                '비목': clean_value(sheet.cell(row=row_num, column=8).value) if has_category else '',
                '비고': clean_value(sheet.cell(row=row_num, column=9).value) if has_category and sheet.max_column >= 9 else clean_value(sheet.cell(row=row_num, column=8).value) if not has_category and sheet.max_column >= 8 else ''
            }

        return record
    except Exception as e:
        raise ValueError(f"행 파싱 오류: {e}")

print("=" * 80)
print("용인시장 업무추진비 데이터 통합 시작 (V3 - 비목 포함)")
print("=" * 80)

# 엑셀 파일 로드
print(f"\n엑셀 파일 로드 중: {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"총 {len(wb.sheetnames)}개 시트 발견")

# 각 시트 처리
for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
    sheet = wb[sheet_name]
    has_category = sheet_idx >= SHEET_WITH_CATEGORY_START

    print(f"\n[{sheet_idx}/{len(wb.sheetnames)}] {sheet_name} 처리 중... (총 {sheet.max_row}행, 비목: {'있음' if has_category else '없음'})")

    # 데이터 시작 행 감지
    format_info = detect_data_start(sheet)
    if not format_info:
        print(f"  [SKIP] 데이터 시작 행을 찾을 수 없음")
        errors.append(f"{sheet_name}: 데이터 시작 행 감지 실패")
        continue

    print(f"  - 데이터 시작 행: {format_info['start_row']}")
    print(f"  - 사용자 컬럼: {'있음' if format_info['has_user_column'] else '없음'}")

    # 데이터 행 처리
    sheet_records = 0
    for row_num in range(format_info['start_row'], sheet.max_row + 1):
        first_cell = clean_value(sheet.cell(row=row_num, column=1).value)
        if not is_numeric(first_cell):
            continue

        try:
            record = parse_row(sheet, row_num, format_info['has_user_column'], has_category)

            if record['사용금액'] <= 0:
                continue

            if not record['사용일시']:
                continue

            all_records.append(record)
            sheet_records += 1

        except Exception as e:
            print(f"  [ERROR] 행 {row_num}: {e}")
            errors.append(f"{sheet_name} 행 {row_num}: {e}")

    print(f"  [OK] {sheet_records}건 추가됨")

print(f"\n{'=' * 80}")
print(f"통합 완료: 총 {len(all_records)}건의 레코드")
print(f"{'=' * 80}")

# CSV 파일로 저장
print(f"\nCSV 파일 저장 중: {OUTPUT_CSV}")
with open(OUTPUT_CSV, 'w', encoding='utf-8-sig', newline='') as f:
    fieldnames = ['번호', '사용자', '사용일시', '사용장소', '집행목적', '대상인원', '사용금액', '결제방법', '비목', '비고']
    writer = csv.DictWriter(f, fieldnames=fieldnames)

    writer.writeheader()
    for record in all_records:
        writer.writerow(record)

print(f"[OK] CSV 파일 저장 완료")

# 통계 출력
if len(all_records) > 0:
    total_amount = sum(r['사용금액'] for r in all_records)

    # 비목 통계 (2024-01 이후만)
    category_records = [r for r in all_records if r['비목'] and r['사용일시'] >= '2024-01']
    if category_records:
        category_stats = {}
        for r in category_records:
            비목 = r['비목']
            if 비목 not in category_stats:
                category_stats[비목] = {'건수': 0, '금액': 0}
            category_stats[비목]['건수'] += 1
            category_stats[비목]['금액'] += r['사용금액']

        print(f"\n{'=' * 80}")
        print("비목 통계 (2024년 1월 이후)")
        print(f"{'=' * 80}")
        for 비목, 통계 in sorted(category_stats.items()):
            print(f"{비목}: {통계['금액']:,}원 ({통계['건수']}건)")

    print(f"\n{'=' * 80}")
    print("통합 데이터 통계")
    print(f"{'=' * 80}")
    print(f"총 건수: {len(all_records):,}건")
    print(f"총 금액: {total_amount:,}원 ({total_amount/100000000:.2f}억원)")
    print(f"평균 금액: {total_amount/len(all_records):,.0f}원")

# 에러 리포트
if errors:
    print(f"\n[WARNING] 발생한 오류 ({len(errors)}건):")
    for error in errors[:10]:
        print(f"  - {error}")
    if len(errors) > 10:
        print(f"  ... 외 {len(errors) - 10}건")

print(f"\n{'=' * 80}")
print("완료!")
print(f"{'=' * 80}")
